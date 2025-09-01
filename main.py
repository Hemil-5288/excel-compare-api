"""
Excel Comparison API (FastAPI)
--------------------------------
Exposes your existing Excel comparison logic as an HTTP API.


Endpoints
- GET /health -> Simple health check
- POST /compare -> Upload two Excel files and (optionally) a JSON sheets_config; returns a compared .xlsx file


Run locally
1) pip install -r requirements.txt
2) uvicorn main:app --host 0.0.0.0 --port 8000 --reload


Example cURL
curl -X POST "http://localhost:8000/compare" \
-F "original_file=@Tax_Report_hemal_patel_28052025 - 4.0.xlsx" \
-F "website_file=@Tax_Report_hemal_patel_21082025.xlsx" \
-F 'sheets_config={"Gain Summary":{"header_row":2,"data_start_row":3},"8938":{"header_row":6,"data_start_row":7},"FBAR":{"header_row":2,"data_start_row":3}}'


Notes
- If sheets_config is omitted, sensible defaults are used.
- Returns an .xlsx file as a binary response with a timestamped filename.
"""

import io
import re
import json
from copy import copy
from datetime import datetime
from typing import Dict, Any, Optional
import pandas as pd
import openpyxl
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import StreamingResponse, JSONResponse
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


from fastapi import FastAPI

app = FastAPI()
app = FastAPI(title="Excel Comparison API", version="1.0.0")


NUMERIC_TOLERANCE = 1

def coerce_numeric(value):
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s == "":
        return None
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    s = s.replace(",", "")
    s = re.sub(r"[^\d\.\-]+", "", s)
    try:
        return float(s)
    except Exception:
        return None


def to_int_or_none(value):
    num = coerce_numeric(value)
    if num is None:
        return None
    try:
        return int(num)
    except Exception:
        return None


def norm_text(value):
    if pd.isna(value) or value is None:
        return ""
    return str(value).strip()


def values_equal(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    oi = to_int_or_none(orig_val)
    wi = to_int_or_none(web_val)
    if oi is not None and wi is not None:
        return abs(wi - oi) <= tol
    return norm_text(orig_val) == norm_text(web_val)


def values_different(orig_val, web_val, tol=NUMERIC_TOLERANCE):
    return not values_equal(orig_val, web_val, tol=tol)


def row_key(row, max_parts=2):
    key_parts = []
    for col in row.index:
        iv = to_int_or_none(row[col])
        if iv is not None:
            key_parts.append(("num", iv))
        else:
            sv = norm_text(row[col])
            if sv != "":
                key_parts.append(("str", sv.lower()))
        if len(key_parts) >= max_parts:
            break
    return tuple(key_parts)


def build_key_index(df, max_parts=2):
    idx = {}
    for i, row in df.iterrows():
        k = row_key(row, max_parts=max_parts)
        idx.setdefault(k, []).append(i)
    return idx


def safe_copy_cell_style(source_cell, target_cell):
    try:
        if source_cell is None or target_cell is None:
            return
        if hasattr(source_cell, "has_style") and source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass


def safe_get_column_width(worksheet, col_letter):
    try:
        if col_letter in worksheet.column_dimensions:
            return worksheet.column_dimensions[col_letter].width
        return None
    except Exception:
        return None


def preprocess_gain_summary(df):
    if "Account Number" not in df.columns:
        return df

    numeric_cols = df.select_dtypes(include="number").columns
    non_numeric_cols = [c for c in df.columns if c not in numeric_cols]

    df_grouped = df.groupby("Account Number", as_index=False).agg(
        {**{col: "first" for col in non_numeric_cols},
         **{col: "sum" for col in numeric_cols}}
    )

    return df_grouped.reset_index(drop=True)


# =============================
# Core compare function (in‑memory)
# =============================

def compare_excel_with_gain_summary_inline(
    original_bytes: bytes,
    website_bytes: bytes,
    sheets_config: Optional[Dict[str, Dict[str, int]]] = None,
) -> bytes:
    if sheets_config is None:
        sheets_config = {
            'Gain Summary': {'header_row': 2, 'data_start_row': 3},
            '8938':         {'header_row': 6, 'data_start_row': 7},
            'FBAR':         {'header_row': 2, 'data_start_row': 3},
        }

    red_fill   = PatternFill(start_color="fbd9d3", end_color="fbd9d3", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    # Load workbooks from memory
    original_wb = openpyxl.load_workbook(io.BytesIO(original_bytes))
    website_wb  = openpyxl.load_workbook(io.BytesIO(website_bytes))
    output_wb   = openpyxl.Workbook()
    if 'Sheet' in output_wb.sheetnames:
        output_wb.remove(output_wb['Sheet'])

    summary_rows = [["Sheet", "Rows Compared", "Cells Different", "Common Rows"]]

    for sheet_name, cfg in sheets_config.items():
        try:
            df_orig = pd.read_excel(io.BytesIO(original_bytes), sheet_name=sheet_name, header=cfg['header_row'] - 1)
            df_web  = pd.read_excel(io.BytesIO(website_bytes),  sheet_name=sheet_name, header=cfg['header_row'] - 1)

            common_cols = [c for c in df_orig.columns if c in df_web.columns]
            if not common_cols:
                summary_rows.append([sheet_name, 0, 0, 0])
                continue

            df_orig = df_orig[common_cols].copy()
            df_web  = df_web[common_cols].copy()

            original_ws = original_wb[sheet_name] if sheet_name in original_wb.sheetnames else None
            website_ws  = website_wb[sheet_name]  if sheet_name in website_wb.sheetnames  else None
            main_ws     = output_wb.create_sheet(title=sheet_name)

            web_key_index = build_key_index(df_web, max_parts=2)
            orig_keys_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

            diff_count        = 0
            rows_compared     = 0
            common_rows_list  = []

            if sheet_name == "Gain Summary":
                headers = []
                for col in common_cols:
                    headers.extend([f"{col} (Original)", f"{col} (Website)", f"{col} (Diff)"])
                for ci, h in enumerate(headers, start=1):
                    main_ws.cell(row=1, column=ci, value=h)
                    if original_ws is not None:
                        header_row_num = cfg['header_row']
                        try:
                            src = original_ws.cell(row=header_row_num, column=((ci - 1) // 3) + 1)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            pass

                for idx_col, _ in enumerate(common_cols, start=1):
                    width = None
                    if original_ws is not None:
                        width = safe_get_column_width(original_ws, get_column_letter(idx_col))
                    for offset in range(3):
                        out_col_letter = get_column_letter((idx_col - 1) * 3 + offset + 1)
                        if width:
                            main_ws.column_dimensions[out_col_letter].width = width

                out_row = 2
                if "Account Number" in df_orig.columns and "Account Number" in df_web.columns:
                    df_orig = df_orig.drop_duplicates(subset=["Account Number"], keep="first").reset_index(drop=True)
                    df_web  = df_web.drop_duplicates(subset=["Account Number"], keep="first").reset_index(drop=True)

                web_key_index = build_key_index(df_web, max_parts=2)
                orig_keys_set = {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}

                for orig_idx, orig_row in df_orig.iterrows():
                    k = row_key(orig_row, max_parts=2)
                    match_indices = web_key_index.get(k, [])
                    match_idx = match_indices[0] if match_indices else None
                    if match_idx is None:
                        for i, col in enumerate(common_cols, start=1):
                            o_val = orig_row[col]
                            c_orig = main_ws.cell(row=out_row, column=(i - 1) * 3 + 1, value=o_val)
                            c_diff = main_ws.cell(row=out_row, column=(i - 1) * 3 + 3, value="Only in Original")
                            c_orig.fill = red_fill
                            c_diff.fill = red_fill
                            if original_ws is not None:
                                src_row = orig_idx + cfg['data_start_row']
                                try:
                                    src_cell = original_ws.cell(row=src_row, column=i)
                                    safe_copy_cell_style(src_cell, c_orig)
                                except Exception:
                                    pass
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1
                        continue

                    web_row = df_web.loc[match_idx]

                    all_same = all(values_equal(orig_row[col], web_row[col]) for col in common_cols)
                    if all_same:
                        common_rows_list.append([orig_row[c] for c in common_cols])
                        continue

                    row_diffs_here = 0
                    for i, col in enumerate(common_cols, start=1):
                        o_val = orig_row[col]
                        w_val = web_row[col]
                        oi = to_int_or_none(o_val)
                        wi = to_int_or_none(w_val)

                        if (oi is not None) and (wi is not None):
                            int_diff = wi - oi
                            show_diff = "" if abs(int_diff) <= NUMERIC_TOLERANCE else int_diff
                            is_diff   = abs(int_diff) > NUMERIC_TOLERANCE
                        else:
                            show_diff = "" if values_equal(o_val, w_val) else "DIFF"
                            is_diff   = not values_equal(o_val, w_val)

                        c_orig = main_ws.cell(row=out_row, column=(i - 1) * 3 + 1, value=o_val)
                        c_web  = main_ws.cell(row=out_row, column=(i - 1) * 3 + 2, value=w_val)
                        c_diff = main_ws.cell(row=out_row, column=(i - 1) * 3 + 3, value=show_diff)

                        if original_ws is not None:
                            src_row = orig_idx + cfg['data_start_row']
                            try:
                                src_cell = original_ws.cell(row=src_row, column=i)
                                safe_copy_cell_style(src_cell, c_orig)
                            except Exception:
                                pass
                        if website_ws is not None:
                            src_row_w = match_idx + cfg['data_start_row']
                            try:
                                src_cell_w = website_ws.cell(row=src_row_w, column=i)
                                safe_copy_cell_style(src_cell_w, c_web)
                            except Exception:
                                pass

                        if is_diff:
                            c_orig.fill = red_fill
                            c_web.fill  = green_fill
                            c_diff.fill = red_fill
                            row_diffs_here += 1

                    out_row += 1
                    rows_compared += 1
                    diff_count += row_diffs_here

                for web_idx, web_row in df_web.iterrows():
                    k = row_key(web_row, max_parts=2)
                    if k in orig_keys_set:
                        continue
                    for i, col in enumerate(common_cols, start=1):
                        c_web  = main_ws.cell(row=out_row, column=(i - 1) * 3 + 2, value=web_row[col])
                        c_diff = main_ws.cell(row=out_row, column=(i - 1) * 3 + 3, value="Only in Website")
                        c_web.fill  = green_fill
                        c_diff.fill = green_fill
                        if website_ws is not None:
                            src_row_w = web_idx + cfg['data_start_row']
                            try:
                                src_cell_w = website_ws.cell(row=src_row_w, column=i)
                                safe_copy_cell_style(src_cell_w, c_web)
                            except Exception:
                                pass
                    out_row += 1
                    rows_compared += 1
                    diff_count += 1

            else:
                for ci, col in enumerate(common_cols, start=1):
                    main_ws.cell(row=1, column=ci, value=col)
                    if original_ws is not None:
                        try:
                            src = original_ws.cell(row=cfg['header_row'], column=ci)
                            safe_copy_cell_style(src, main_ws.cell(row=1, column=ci))
                        except Exception:
                            pass

                out_row = 2
                for orig_idx, orig_row in df_orig.iterrows():
                    k = row_key(orig_row, max_parts=2)
                    match_indices = web_key_index.get(k, [])
                    match_idx = match_indices[0] if match_indices else None

                    if match_idx is None:
                        for i, col in enumerate(common_cols, start=1):
                            c = main_ws.cell(row=out_row, column=i, value=orig_row[col])
                            c.fill = red_fill
                            if original_ws is not None:
                                src_row = orig_idx + cfg['data_start_row']
                                try:
                                    src_cell = original_ws.cell(row=src_row, column=i)
                                    safe_copy_cell_style(src_cell, c)
                                except Exception:
                                    pass
                        out_row += 1
                        rows_compared += 1
                        diff_count += 1
                        continue

                    web_row = df_web.loc[match_idx]
                    if all(values_equal(orig_row[col], web_row[col]) for col in common_cols):
                        common_rows_list.append([orig_row[c] for c in common_cols])
                        continue

                    row_diffs_here = 0
                    for i, col in enumerate(common_cols, start=1):
                        o_val = orig_row[col]
                        w_val = web_row[col]
                        c = main_ws.cell(row=out_row, column=i, value=o_val)

                        if values_different(o_val, w_val):
                            c.fill = red_fill
                            row_diffs_here += 1

                        if original_ws is not None:
                            src_row = orig_idx + cfg['data_start_row']
                            try:
                                src_cell = original_ws.cell(row=src_row, column=i)
                                safe_copy_cell_style(src_cell, c)
                            except Exception:
                                pass

                    out_row += 1
                    rows_compared += 1
                    diff_count += row_diffs_here

                for web_idx, web_row in df_web.iterrows():
                    k = row_key(web_row, max_parts=2)
                    if k in {row_key(r, max_parts=2) for _, r in df_orig.iterrows()}:
                        continue
                    for i, col in enumerate(common_cols, start=1):
                        c = main_ws.cell(row=out_row, column=i, value=web_row[col])
                        c.fill = green_fill
                        if website_ws is not None:
                            src_row_w = web_idx + cfg['data_start_row']
                            try:
                                src_cell_w = website_ws.cell(row=src_row_w, column=i)
                                safe_copy_cell_style(src_cell_w, c)
                            except Exception:
                                pass
                    out_row += 1
                    rows_compared += 1
                    diff_count += 1

            # Common rows sheet
            common_title = f"{sheet_name} Common Rows"
            base_title = common_title
            idx_count = 1
            while common_title in output_wb.sheetnames:
                common_title = f"{base_title}_{idx_count}"
                idx_count += 1
            ws_common = output_wb.create_sheet(title=common_title)

            for ci, col in enumerate(common_cols, start=1):
                cell = ws_common.cell(row=1, column=ci, value=col)
                if original_ws is not None:
                    try:
                        src = original_ws.cell(row=cfg['header_row'], column=ci)
                        safe_copy_cell_style(src, cell)
                    except Exception:
                        pass

            r_out = 2
            for rowvals in common_rows_list:
                for ci, val in enumerate(rowvals, start=1):
                    ws_common.cell(row=r_out, column=ci, value=val)
                r_out += 1

            for ws_auto in (main_ws, ws_common):
                for col in ws_auto.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    ws_auto.column_dimensions[col_letter].width = min(max_len + 2, 50)

            summary_rows.append([sheet_name, rows_compared, diff_count, len(common_rows_list)])

        except Exception as e:
            summary_rows.append([sheet_name, 0, f"ERROR: {e}", 0])

    ws_summary = output_wb.create_sheet("Summary", index=0)
    ws_summary.append(["Excel Comparison Report"])
    ws_summary.append([f"Generated On:  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    for row in summary_rows:
        ws_summary.append(row)

    out_bytes = io.BytesIO()
    output_wb.save(out_bytes)
    out_bytes.seek(0)
    return out_bytes.getvalue()


# =============================
# API Endpoints
# =============================

@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.utcnow().isoformat() + "Z"}

@app.post("/compare")
async def compare_endpoint(
    original_file: UploadFile = File(..., description="Original Excel (.xlsx)"),
    website_file: UploadFile = File(..., description="Website Excel (.xlsx)"),
    sheets_config: Optional[str] = Form(
        default=None,
        description="Optional JSON mapping of sheet -> {header_row, data_start_row}",
    ),
):
    try:
        parsed_config = None
        if sheets_config and sheets_config.strip():
          try:
            parsed_config = json.loads(sheets_config.strip())
          except Exception:
            parsed_config = None 

        original_bytes = await original_file.read()
        website_bytes = await website_file.read()

        result_bytes = compare_excel_with_gain_summary_inline(
            original_bytes, website_bytes, sheets_config=parsed_config
        )

        filename = f"comparison_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            io.BytesIO(result_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        return {"error": str(e)}
